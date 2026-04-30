"""Seed-Import: Excel-Spieltagsdaten -> Supabase Cloud-DB.

Liest Windmill_Poker_results_test.xlsx (Sheet "2025") und schreibt Spieler,
Spieltage, Teilnahmen und Auszahlungen idempotent in die Cloud-DB.

Datenmodell:
    - Rechte Seite (Cols AP..AR): pro Spieltag Tabelle "Spieler | Betrag | Plzg".
      Quelle für played_on (aus ST-Header) und payout pro Spieler.
    - Linke Seite (anwesend-Zeilen): wer war anwesend - nur dafür gebraucht,
      damit Spieler mit 0EUR-Spieltagen in der Teilnahme-Statistik mitzählen.
    - Match links<->rechts per Reihenfolge (ST1 = 1. Block, ..., ST10 = 10.).

Usage:
    python scripts/seed_import.py --dry-run
    python scripts/seed_import.py
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import Client, create_client

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "Windmill_Poker_results_test.xlsx"
SHEET_NAME = "2025"

# Spieler-Spalten Sheet "2025" (1-indexed). Flag-Col: "1" wenn anwesend.
PLAYER_COLS: list[tuple[str, int]] = [
    ("Werner", 9), ("Peter", 11), ("Rainer", 13), ("Friedl", 15),
    ("Torben", 17), ("Jens", 19), ("Jörg", 21), ("Jochen", 23),
    ("Martin", 25), ("Frank", 27), ("Ciano", 29),
]

# Anwesend-Marker-Zeilen, in Reihenfolge ST1..ST10.
ANWESEND_ROWS = [14, 24, 34, 44, 54, 66, 76, 86, 96, 108]

# Rechte-Seite Spalten der Auszahlungs-Tabellen.
COL_PLAYER = 42  # AP
COL_BETRAG = 43  # AQ

# Header-Pattern: "ST1 (16.01.) - 6 Spieler" oder "ST10 (15.01.26) - 8 Spieler"
HEADER_RE = re.compile(r"^ST(\d+)\s*\((\d{2})\.(\d{2})\.(\d{2})?\.?\)")


@dataclass
class GameDay:
    played_on: date
    attendees: list[str]                # Namen aus links-anwesend
    payouts: dict[str, Decimal] = field(default_factory=dict)  # name -> EUR


# =============================================================================
# Parsing
# =============================================================================

def parse_attendees_left(ws) -> list[list[str]]:
    """Liefert pro ST-Index die Anwesenheitsliste aus den linken anwesend-Zeilen."""
    result = []
    for r in ANWESEND_ROWS:
        attendees = [
            name for (name, flag_c) in PLAYER_COLS
            if ws.cell(row=r, column=flag_c).value == 1
        ]
        result.append(attendees)
    return result


def parse_payout_blocks_right(ws) -> list[GameDay]:
    """Findet alle ST-Header rechts, parst Datum + Spieler/Betrag-Tabelle."""
    blocks: list[GameDay] = []
    last_row = ws.max_row

    r = 1
    while r <= last_row:
        v = ws.cell(row=r, column=COL_PLAYER).value
        if isinstance(v, str) and (m := HEADER_RE.match(v)):
            day = int(m.group(2))
            month = int(m.group(3))
            year_2digit = m.group(4)
            year = 2000 + int(year_2digit) if year_2digit else 2025
            played_on = date(year, month, day)

            # Skip header + "Spieler|Betrag|Plzg" header row -> R+2 ist erste Datenzeile.
            payouts: dict[str, Decimal] = {}
            i = r + 2
            while i <= last_row:
                name = ws.cell(row=i, column=COL_PLAYER).value
                betrag = ws.cell(row=i, column=COL_BETRAG).value
                if name is None:
                    break
                if name == "SUMME":
                    break
                if isinstance(name, str):
                    cleaned_name = name.strip()
                    cents = _to_cents(betrag)
                    payouts[cleaned_name] = Decimal(cents) / 100
                i += 1

            blocks.append(GameDay(played_on=played_on, attendees=[], payouts=payouts))
            r = i + 1
        else:
            r += 1

    return blocks


def _to_cents(v) -> int:
    if v is None or v == "":
        return 0
    return int(round(float(v) * 100))


# =============================================================================
# Validierung
# =============================================================================

def validate(blocks: list[GameDay]) -> None:
    """Sanity: 10 Blöcke, eindeutige Daten, jeder Spieler in payouts auch anwesend."""
    if len(blocks) != 10:
        raise ValueError(f"Erwarte 10 Spieltage, gefunden: {len(blocks)}")

    seen_dates = set()
    for gd in blocks:
        if gd.played_on in seen_dates:
            raise ValueError(f"Doppeltes Datum: {gd.played_on}")
        seen_dates.add(gd.played_on)

        # Spieler in payouts müssen in attendees stehen (sonst Datenfehler).
        attendee_set = set(gd.attendees)
        for name in gd.payouts:
            if name not in attendee_set:
                raise ValueError(
                    f"{gd.played_on}: '{name}' hat Auszahlung, ist aber laut "
                    f"linker Seite nicht anwesend. Anwesend: {sorted(attendee_set)}"
                )


# =============================================================================
# Cloud-Insert
# =============================================================================

def upsert_all(supabase: Client, game_days: list[GameDay]) -> None:
    all_names = sorted({n for gd in game_days for n in gd.attendees})

    print(f"[1/4] upsert {len(all_names)} players ...")
    supabase.table("players").upsert(
        [{"name": n, "is_active": True} for n in all_names],
        on_conflict="name",
    ).execute()
    sel = supabase.table("players").select("id,name").in_("name", all_names).execute()
    name_to_id = {row["name"]: row["id"] for row in sel.data}

    print(f"[2/4] upsert {len(game_days)} game_days ...")
    supabase.table("game_days").upsert(
        [{"played_on": gd.played_on.isoformat(), "is_closed": True}
         for gd in game_days],
        on_conflict="played_on",
    ).execute()
    sel = supabase.table("game_days").select("id,played_on").in_(
        "played_on", [gd.played_on.isoformat() for gd in game_days]
    ).execute()
    date_to_gd_id = {row["played_on"]: row["id"] for row in sel.data}

    print(f"[3/4] upsert attendances ...")
    att_rows = []
    for gd in game_days:
        gd_id = date_to_gd_id[gd.played_on.isoformat()]
        for name in gd.attendees:
            att_rows.append({"game_day_id": gd_id, "player_id": name_to_id[name]})
    supabase.table("attendances").upsert(
        att_rows, on_conflict="game_day_id,player_id"
    ).execute()
    print(f"      {len(att_rows)} attendances")

    print(f"[4/4] upsert round_results (round_number=1, Gesamt-Betrag) ...")
    rr_rows = []
    for gd in game_days:
        gd_id = date_to_gd_id[gd.played_on.isoformat()]
        for name, betrag in gd.payouts.items():
            if betrag == 0:
                continue
            rr_rows.append({
                "game_day_id": gd_id,
                "player_id": name_to_id[name],
                "round_number": 1,
                "payout": float(betrag),
            })
    supabase.table("round_results").upsert(
        rr_rows, on_conflict="game_day_id,player_id,round_number"
    ).execute()
    print(f"      {len(rr_rows)} round_results")


# =============================================================================
# Output
# =============================================================================

def print_summary(game_days: list[GameDay]) -> None:
    print()
    print("=" * 70)
    print(f"{'Datum':<12} {'N':>3}  {'Topf':>8}  Spieler -> Betrag")
    print("=" * 70)
    for gd in game_days:
        topf = sum(gd.payouts.values())
        print(f"{gd.played_on}  {len(gd.attendees):>3}  {topf:>7}EUR")
        for name in gd.attendees:
            betrag = gd.payouts.get(name, Decimal(0))
            mark = " " if betrag > 0 else "*"
            print(f"  {mark} {name:<8}  {betrag:>7}EUR")
        print()

    # Gesamt-Rangliste
    totals: dict[str, Decimal] = {}
    plays: dict[str, int] = {}
    for gd in game_days:
        for name in gd.attendees:
            totals[name] = totals.get(name, Decimal(0)) + gd.payouts.get(name, Decimal(0))
            plays[name] = plays.get(name, 0) + 1

    print("=" * 70)
    print("Rangliste (zur Verifikation gegen Excel R102+)")
    print("=" * 70)
    print(f"{'Spieler':<10} {'Total':>9}  {'Spt':>3}  {'EUR/Spt':>8}")
    for name in sorted(totals, key=lambda n: -totals[n]):
        avg = totals[name] / plays[name]
        print(f"{name:<10} {totals[name]:>8}EUR  {plays[name]:>3}  {avg:>7}EUR")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print(f"Reading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    attendees_per_st = parse_attendees_left(ws)
    blocks = parse_payout_blocks_right(ws)

    if len(attendees_per_st) != len(blocks):
        print(
            f"FEHLER: linke Seite hat {len(attendees_per_st)} Bloecke, "
            f"rechte Seite {len(blocks)}.",
            file=sys.stderr,
        )
        return 2

    for gd, attendees in zip(blocks, attendees_per_st):
        gd.attendees = attendees

    validate(blocks)
    print(f"OK: {len(blocks)} Spieltage geparst.")
    print_summary(blocks)

    if args.dry_run:
        print("\n[dry-run] kein DB-Write.")
        return 0

    load_dotenv(REPO_ROOT / ".env.local")
    url = os.environ.get("SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not service_key:
        print(
            "\nFEHLER: SUPABASE_URL und SUPABASE_SERVICE_ROLE_KEY muessen in "
            ".env.local stehen. Den Service-Role-Key gibt's im Dashboard "
            "unter Settings > API.",
            file=sys.stderr,
        )
        return 2

    sb = create_client(url, service_key)
    upsert_all(sb, blocks)
    print("\nFertig. Verifikation: SELECT * FROM moneylist;")
    return 0


if __name__ == "__main__":
    sys.exit(main())
