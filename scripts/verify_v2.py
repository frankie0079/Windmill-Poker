"""Cross-Check: jede Zelle Auszahlungen-Sheet vs. DB.

Vergleicht für jeden Spieler × jeden Spieltag:
  - Excel-Cellwert (None / Number)
  - DB: hat Attendance (ja/nein) + payout (Σ round_results)

Bricht NICHT ab, sammelt alle Diffs und gibt Liste am Ende.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "Windmill_Poker_results_test.xlsx"

ROW_DATES = 14
ROW_PLAYER_FIRST = 17
ROW_PLAYER_LAST = 27
COL_ST_FIRST = 2
COL_ST_LAST = 12  # ST1..ST11; ST12 (C13) skipped


def parse_date_cell(v):
    from datetime import date
    if hasattr(v, "year"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().rstrip(".")
    parts = s.split(".")
    if len(parts) == 2:
        d, m = int(parts[0]), int(parts[1])
        return date(2025, m, d).strftime("%Y-%m-%d")
    if len(parts) == 3:
        d, m, y = parts
        year = int(y)
        if year < 100:
            year += 2000
        return date(year, int(m), int(d)).strftime("%Y-%m-%d")
    raise ValueError(f"Cannot parse date: {v!r}")


def main() -> int:
    print(f"Reading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Auszahlungen"]

    # ST-Daten (ST1..ST11)
    st_cols = list(range(COL_ST_FIRST, COL_ST_LAST + 1))
    st_dates = [parse_date_cell(ws.cell(row=ROW_DATES, column=c).value) for c in st_cols]
    st_idx = list(range(1, len(st_cols) + 1))

    # Excel-Daten: {(name, st_idx): value or None}
    excel = {}
    for r in range(ROW_PLAYER_FIRST, ROW_PLAYER_LAST + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        for i, c in zip(st_idx, st_cols):
            v = ws.cell(row=r, column=c).value
            excel[(name, i)] = float(v) if v is not None else None

    # DB-Daten holen
    load_dotenv(REPO_ROOT / ".env.local")
    sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])

    pl = {row["id"]: row["name"] for row in sb.table("players").select("id,name").execute().data}
    name_to_id = {v: k for k, v in pl.items()}

    gd_data = sb.table("game_days").select("id,played_on").execute().data
    date_to_gd_id = {row["played_on"]: row["id"] for row in gd_data}

    att_data = sb.table("attendances").select("game_day_id,player_id").execute().data
    att_set = {(r["game_day_id"], r["player_id"]) for r in att_data}

    rr_data = sb.table("round_results").select("game_day_id,player_id,payout").execute().data
    payout_db = {}
    for r in rr_data:
        key = (r["game_day_id"], r["player_id"])
        payout_db[key] = payout_db.get(key, 0) + float(r["payout"])

    # Vergleich pro Zelle
    print()
    print("=" * 90)
    print(f"{'Spieler':<10} " + " ".join(f"ST{i:>2}" for i in st_idx))
    print("=" * 90)
    diffs = []
    for r in range(ROW_PLAYER_FIRST, ROW_PLAYER_LAST + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        pid = name_to_id.get(name)
        line = f"{name:<10} "
        for i, st_date in zip(st_idx, st_dates):
            xv = excel[(name, i)]
            gd_id = date_to_gd_id.get(st_date)
            if not gd_id or not pid:
                line += "  ?  "
                continue
            in_att = (gd_id, pid) in att_set
            db_pay = payout_db.get((gd_id, pid), 0)

            # Excel: None -> nicht angetreten. Nicht-None -> angetreten mit Wert.
            if xv is None:
                if in_att or db_pay > 0:
                    diffs.append((name, i, st_date, xv, in_att, db_pay))
                    line += " ✗   "
                else:
                    line += " ·   "  # both empty
            else:
                # Excel hat Wert. DB sollte Attendance haben + payout=xv (oder 0 wenn 0).
                if not in_att:
                    # Excel sagt teilgenommen, DB sagt nicht
                    diffs.append((name, i, st_date, xv, in_att, db_pay))
                    line += "MISS "
                elif abs(db_pay - xv) > 0.01:
                    diffs.append((name, i, st_date, xv, in_att, db_pay))
                    line += f" ✗   "
                else:
                    line += f"{int(xv):>3}  " if xv == int(xv) else f"{xv:>5.1f}"
        print(line)

    print()
    print("=" * 90)
    if diffs:
        print(f"FEHLER: {len(diffs)} Diffs")
        for name, st, dt, xv, in_att, db_pay in diffs:
            print(f"  ST{st} ({dt}) {name}: Excel={xv} | DB attendance={in_att} payout={db_pay}")
    else:
        print(f"OK: alle {len(excel)} Zellen identisch zur DB.")

    # Zusatzcheck: Σ Moneylist + Teilnahmen
    print()
    print("=== Σ und Teilnahmen-Counts ===")
    for r in range(ROW_PLAYER_FIRST, ROW_PLAYER_LAST + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        excel_sum = ws.cell(row=r, column=14).value or 0
        excel_n = ws.cell(row=r, column=16).value or 0
        pid = name_to_id.get(name)
        db_sum = sum(p for k, p in payout_db.items() if k[1] == pid)
        db_n = sum(1 for k in att_set if k[1] == pid)
        ok_sum = "✓" if abs(float(excel_sum) - db_sum) < 0.01 else "✗"
        ok_n = "✓" if int(excel_n) == db_n else "✗"
        print(f"  {ok_sum}{ok_n} {name:<10} Excel: Σ={excel_sum} N={excel_n} | DB: Σ={db_sum:.2f} N={db_n}")

    return 0 if not diffs else 1


if __name__ == "__main__":
    sys.exit(main())
