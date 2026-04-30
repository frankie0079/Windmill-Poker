"""Re-Seed v2 (autoritativ): Liest Auszahlungen-Sheet als Source of Truth.

Sheet "Auszahlungen" der Windmill_Poker_results_test.xlsx hat in R13/R14/R15/R16
die ST-Header (ST1..ST12), Daten, Spielerzahl, Topf. Spalten C2..C13 sind ST1..ST12.
Zeilen R17..R27 sind die 11 Spieler mit Auszahlungen pro ST + Σ Moneylist + Teilnahmen.

ST12 (05.03.26) ist in der Excel ein Platzhalter (alle Werte 0, noch nicht eingegeben)
und wird hier ausgelassen.

ST11 (19.02.26) war 8-Spieler-Spieltag. Die Teilnahmen pro Spieler ergeben sich
über die Differenz: Excel-Teilnahmen-Spalte minus Anwesenheits-Count aus Sheet
"2025" (R14..R108 anwesend-Zeilen für ST1..ST10).

Reset-Modus: löscht round_results, attendances, next_game_planning und nicht-genutzte
game_days, dann fresh insert. Idempotent.

Außerdem: setzt next_game_date=07.05.26 auf ST11 + 8 Teilnehmer + 3 Warteliste
(Mockup-Vorgabe Deckblatt).
"""
from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "Windmill_Poker_results_test.xlsx"

# Spieler-Spalten in Sheet "2025" (1-indexed Flag-Col).
ANW_PLAYERS = [
    ("Werner", 9), ("Peter", 11), ("Rainer", 13), ("Friedl", 15),
    ("Torben", 17), ("Jens", 19), ("Jörg", 21), ("Jochen", 23),
    ("Martin", 25), ("Frank", 27), ("Ciano", 29),
]
ANW_ROWS_2025 = [14, 24, 34, 44, 54, 66, 76, 86, 96, 108]  # ST1..ST10

# Auszahlungen-Sheet Spalten
COL_ST_FIRST = 2   # C2 = ST1
COL_ST_LAST = 13   # C13 = ST12
COL_TOTAL = 14
COL_AVG = 15
COL_COUNT = 16

ROW_ST_HEADERS = 13  # ST1..ST12
ROW_DATES = 14
ROW_SPIELER_N = 15
ROW_TOPF = 16
ROW_PLAYER_FIRST = 17
ROW_PLAYER_LAST = 27

# Mockup-Vorgabe: Nächster Spieltag = 07.05.26 (= über ST11 hinaus geplant).
NEXT_GAME_DATE = "2026-05-07"
NEXT_PARTICIPANTS = ["Frank", "Peter", "Friedl", "Jörg", "Rainer", "Jochen", "Ciano", "Torben"]
NEXT_WAITLIST = [("Werner", 1), ("Jens", 2), ("Martin", 3)]


def parse_date_cell(v) -> str:
    """Excel-Datum -> ISO 'YYYY-MM-DD'. Default-Jahr 2025 für strings ohne Jahr."""
    if hasattr(v, "year"):  # datetime
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().rstrip(".")
    parts = s.split(".")
    if len(parts) == 2:
        d, m = int(parts[0]), int(parts[1])
        return date(2025, m, d).strftime("%Y-%m-%d")
    if len(parts) == 3:
        d, m, y = parts
        year = int(y)
        if year < 100:
            year += 2000
        return date(year, int(m), int(d)).strftime("%Y-%m-%d")
    raise ValueError(f"Cannot parse date: {v!r}")


def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_a = wb["Auszahlungen"]
    ws_2025 = wb["2025"]

    # ST-Liste aus Auszahlungen lesen
    sts = []
    for c in range(COL_ST_FIRST, COL_ST_LAST + 1):
        date_iso = parse_date_cell(ws_a.cell(row=ROW_DATES, column=c).value)
        n_spieler = ws_a.cell(row=ROW_SPIELER_N, column=c).value
        topf = ws_a.cell(row=ROW_TOPF, column=c).value
        sts.append({"col": c, "date": date_iso, "n": n_spieler, "topf": topf})
    # Letzten ST (ST12) als Platzhalter erkennen: alle Auszahlungen 0
    st12 = sts[-1]
    all_zero = True
    for r in range(ROW_PLAYER_FIRST, ROW_PLAYER_LAST + 1):
        v = ws_a.cell(row=r, column=st12["col"]).value
        if v is not None and v != 0:
            all_zero = False
            break
    if all_zero:
        print(f"  ST12 ({st12['date']}) übersprungen (keine Daten).")
        sts = sts[:-1]

    # Spieler-Daten aus Auszahlungen lesen
    players = {}
    for r in range(ROW_PLAYER_FIRST, ROW_PLAYER_LAST + 1):
        name = ws_a.cell(row=r, column=1).value
        if not name:
            continue
        teilnahmen = ws_a.cell(row=r, column=COL_COUNT).value or 0
        per_st = {}
        for st in sts:
            v = ws_a.cell(row=r, column=st["col"]).value
            if v is not None:
                per_st[st["date"]] = float(v)
        players[name] = {"teilnahmen": int(teilnahmen), "per_st": per_st}

    # Sheet-2025-Anwesenheit für ST1..ST10
    anw_2025 = {name: set() for name, _ in ANW_PLAYERS}
    for st_idx, r in enumerate(ANW_ROWS_2025):
        st_date = sts[st_idx]["date"]
        for name, col in ANW_PLAYERS:
            if ws_2025.cell(row=r, column=col).value == 1:
                anw_2025[name].add(st_date)

    return sts, players, anw_2025


def derive_attendance(sts, players, anw_2025):
    """Liefert {date: set(player_names)} für ST1..ST11.

    Regel:
    - ST1..ST10: anwesend laut Sheet-2025-Matrix
    - ST11: Spieler hat ST11 attended ↔ (Excel-Teilnahmen − len(anw_2025-Set)) > 0
    """
    attendance = {st["date"]: set() for st in sts}

    # ST1..ST10
    for st in sts[:10]:
        for name, _ in ANW_PLAYERS:
            if st["date"] in anw_2025.get(name, set()):
                attendance[st["date"]].add(name)

    # ST11
    if len(sts) >= 11:
        st11 = sts[10]
        for name, info in players.items():
            n_left = len(anw_2025.get(name, set()))
            n_total = info["teilnahmen"]
            if n_total > n_left:
                attendance[st11["date"]].add(name)

    # Sanity: Teilnehmer-Anzahl pro ST muss zur Spielerzahl passen.
    for st in sts:
        actual = len(attendance[st["date"]])
        expected = int(st["n"])
        if actual != expected:
            print(f"  WARN {st['date']}: {actual} attendees, Excel sagt {expected}")

    return attendance


def reset_db(sb: Client):
    print("[reset] DELETE round_results, attendances, next_game_planning, game_days ...")
    sb.table("round_results").delete().neq("payout", -999999).execute()
    sb.table("attendances").delete().neq("game_day_id", "00000000-0000-0000-0000-000000000000").execute()
    sb.table("next_game_planning").delete().neq("game_day_id", "00000000-0000-0000-0000-000000000000").execute()
    sb.table("game_days").delete().neq("played_on", "1900-01-01").execute()


def main() -> int:
    print(f"Reading {EXCEL_PATH} ...")
    sts, players, anw_2025 = load_excel()
    print(f"  {len(sts)} Spieltage geladen, {len(players)} Spieler.")

    attendance = derive_attendance(sts, players, anw_2025)

    load_dotenv(REPO_ROOT / ".env.local")
    sb = create_client(
        os.environ["SUPABASE_URL"],
        os.environ["SUPABASE_SERVICE_ROLE_KEY"],
    )

    reset_db(sb)

    # 1) Players (alle aus der Liste, auch Martin mit 0 Teilnahmen)
    print(f"[1/4] upsert {len(players)} players ...")
    sb.table("players").upsert(
        [{"name": name, "is_active": True} for name in players],
        on_conflict="name",
    ).execute()
    sel = sb.table("players").select("id,name").in_(
        "name", list(players.keys())
    ).execute()
    name_to_id = {row["name"]: row["id"] for row in sel.data}

    # 2) game_days. ST11 bekommt next_game_date=07.05.26. Alle is_closed=true (Daten erfasst).
    print(f"[2/4] insert {len(sts)} game_days ...")
    gd_payload = []
    for i, st in enumerate(sts):
        is_st11 = i == 10
        gd_payload.append({
            "played_on": st["date"],
            "is_closed": True,
            "next_game_date": NEXT_GAME_DATE if is_st11 else None,
        })
    sb.table("game_days").insert(gd_payload).execute()
    sel = sb.table("game_days").select("id,played_on").in_(
        "played_on", [st["date"] for st in sts]
    ).execute()
    date_to_gd_id = {row["played_on"]: row["id"] for row in sel.data}

    # 3) attendances
    print(f"[3/4] insert attendances ...")
    att_rows = []
    for st in sts:
        gd_id = date_to_gd_id[st["date"]]
        for name in attendance[st["date"]]:
            att_rows.append({
                "game_day_id": gd_id,
                "player_id": name_to_id[name],
            })
    sb.table("attendances").insert(att_rows).execute()
    print(f"      {len(att_rows)} attendances")

    # 4) round_results (round_number=1, Gesamt-Betrag pro ST)
    print(f"[4/4] insert round_results ...")
    rr_rows = []
    for st in sts:
        gd_id = date_to_gd_id[st["date"]]
        for name in attendance[st["date"]]:
            payout = players[name]["per_st"].get(st["date"], 0)
            if payout == 0:
                continue
            rr_rows.append({
                "game_day_id": gd_id,
                "player_id": name_to_id[name],
                "round_number": 1,
                "payout": payout,
            })
    sb.table("round_results").insert(rr_rows).execute()
    print(f"      {len(rr_rows)} round_results (mit Betrag > 0)")

    # 5) next_game_planning für ST11 mit den 8+3 aus Mockup
    st11_id = date_to_gd_id[sts[10]["date"]]
    plan_rows = []
    for name in NEXT_PARTICIPANTS:
        plan_rows.append({
            "game_day_id": st11_id,
            "player_id": name_to_id[name],
            "role": "participant",
            "status": "confirmed",
            "waitlist_rank": None,
        })
    for name, rank in NEXT_WAITLIST:
        plan_rows.append({
            "game_day_id": st11_id,
            "player_id": name_to_id[name],
            "role": "waitlist",
            "status": None,
            "waitlist_rank": rank,
        })
    sb.table("next_game_planning").insert(plan_rows).execute()
    print(f"[+] next_game_planning für ST11 -> {NEXT_GAME_DATE}: 8 Teilnehmer + 3 Warteliste")

    # Verifikation
    print()
    print("=== Verifikation moneylist gegen Excel-Rangliste ===")
    ml = sb.table("moneylist").select("name,total_winnings,game_days_played,avg_per_gameday").execute()
    for row in ml.data:
        excel = players.get(row["name"])
        excel_total = excel["per_st"].get if excel else None
        # Excel hat C14 = Σ Moneylist
        # Wir koennen das hier nicht direkt aufrufen — vergleichen wir per Sum-of-cells.
        excel_sum = sum(excel["per_st"].values()) if excel else None
        ok = "✓" if abs((excel_sum or 0) - float(row["total_winnings"])) < 0.01 else "✗"
        print(f"  {ok} {row['name']:<10} DB={row['total_winnings']:>7.2f}  cells_sum={excel_sum:>7.2f}  ST={row['game_days_played']}  avg={row['avg_per_gameday']:>6.2f}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
